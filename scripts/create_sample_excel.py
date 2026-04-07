"""生成一份模拟的甲方资产包Excel，用于系统测试"""

import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "资产包清单"

# 表头
headers = ["序号", "车型", "VIN码", "首次登记日期", "GPS状态", "是否脱保", "是否过户", "债权本金(元)", "买断价(元)"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

# 模拟数据（VIN为示例格式，非真实车辆）
data = [
    [1, "2019 丰田凯美瑞 2.0G 豪华版", "LFMA12GL4K0012345", "2019-03-15", "在线", "否", "否", 120000, 48000],
    [2, "2020 本田雅阁 1.5T 精英版", "LHGCV2F35LA001234", "2020-06-20", "在线", "否", "否", 150000, 60000],
    [3, "2018 大众帕萨特 330TSI 精英版", "LSVNL4DC5JN012345", "2018-09-10", "离线", "是", "否", 100000, 35000],
    [4, "2021 宝马3系 325Li M运动套装", "LBV5S3109MSK01234", "2021-01-08", "在线", "否", "否", 280000, 140000],
    [5, "2017 别克君威 20T 精英型", "LSGKE52H0HB012345", "2017-11-25", "离线", "是", "是", 80000, 20000],
    [6, "2020 日产天籁 2.0L XL 舒适版", "LGBH12E05LY012345", "2020-04-18", "在线", "否", "否", 130000, 52000],
    [7, "2019 奥迪A4L 40TFSI 时尚型", "LFV3A28K4K3012345", "2019-07-22", "在线", "否", "否", 200000, 90000],
    [8, "2016 现代名图 1.8L 自动智能型", "LBEHDAFC7GY012345", "2016-05-30", "离线", "是", "否", 60000, 15000],
    [9, "2021 吉利星瑞 2.0TD 尊贵型", "L6T79Y2E3MA012345", "2021-08-12", "在线", "否", "否", 100000, 45000],
    [10, "2018 福特蒙迪欧 EcoBoost180 豪华型", "LVSHFFAL5JF012345", "2018-12-05", "离线", "否", "是", 110000, 30000],
]

for row_data in data:
    ws.append(row_data)

# 调整列宽
col_widths = [6, 35, 20, 15, 8, 8, 8, 15, 12]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

output_path = os.path.join(os.path.dirname(__file__), "..", "backend", "data", "sample_asset_package.xlsx")
wb.save(output_path)
print(f"示例Excel已生成: {output_path}")
